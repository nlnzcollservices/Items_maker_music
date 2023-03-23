import re
import os
import sys
import PySimpleGUI as sg
import urllib3
import random
from datetime import datetime as dt
from alma_tools_mod import AlmaTools
from openpyxl import load_workbook
from time import sleep

# to disable warnings verify=False
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# setting arrival date
arrival_date = dt.now().strftime("%Y-%m-%dZ")


def my_gui(values=None):
	"""
	This function runs a form in a separate window and collects values.
	Returns:
	    values (dict) - dictionary of form values
	"""
	if not values:
	    spreadsheet_name = "test_sheet.xlsx"

	form = sg.FlexForm("xlsx submit form")
	schemes = ["Green", "DarkGreen", "GreenMono"]

	r = random.choice(schemes)
	sg.theme(r)

	layout = [
	    [
	        sg.Text(
	            "This app is making items with barcodes from xlsx file.",
	            font=("Courier",13, "bold")
	        )
	    ],
	    [
	        sg.Text(
	            "Please select your spreadsheet.", font=("Courier",13,"roman")
	        )
	    ],
	    [
	        sg.Text("browse results", size=(30, 1), font = ("Helvetica",9, "italic")),
	        sg.FileBrowse("FileBrowse", key="spreadsheet_name"),
	    ],
	    [
	        sg.Checkbox(
	            "Try in SANDBOX?",
	            default=False,
	            key="if_sb",
	            font=("Tahoma", 12),
	        )
	    ],
	    [sg.Button("Run!"), sg.Button("Exit")],
	]

	window = sg.Window(
	    f'Item maker ("{r}" theme)',
	    layout,
	    default_element_size=(35, 2),
	)

	while True:
	    event, values = window.read()
	    if event in (sg.WIN_CLOSED, "Exit"):
	        window.close()
	        break
	    elif event == "Run!":
	        return values, window, event  # Return three values



def make_item(mms_id, holding_id=None, location=None, copy_id=None, barcode=None, key=None):
	"""Thid function creates item
	Parameters:
		mms_id(str) - Alms mms id
		holding_id(str) - Alma  holding id (optional, defaults to None)
		location(str) - location status_code (optional, defaults to None)
		copy_id(str) - identification number (optional, defaults to None)
		barcode(str) - barcode (optional, defaults to None)
		key(str) - "prod"  or "sb" for passing to AlmaTools (optional, defaults to None)
	Returns:
		flag(bool) - True if item created and False if Failed
	"""
	my_alma = AlmaTools(key)
	if location is None and holding_id is None:
		raise ValueError("Either a location or holding_id must be provided.")
	else:
		if not holding_id:
			my_alma.get_holdings(mms_id)
			holdings = re.findall(r'<holding_id>(.*?)</holding_id>', my_alma.xml_response_data)
		else:
			holdings = [holding_id]
		print("here")
		for holding_id in holdings:
			print(holding_id)
			my_alma.get_holding(mms_id, holding_id)
			print(my_alma.xml_response_data)
			if location in my_alma.xml_response_data:
				sg.Print("MMS Id: ", mms_id)
				sg.Print("Holding id: ", holding_id)
				item_data = """
					<item>
					<holding_data>
						<holding_id></holding_id>
						<copy_id></copy_id>
					</holding_data>
					<item_data>
						<is_magnetic>false</is_magnetic>
						<barcode></barcode>
						<physical_material_type desc="Music Score">SCORE</physical_material_type>
						<policy desc="Music Hire">HIRE</policy>
						<enumeration_a></enumeration_a>
						<enumeration_b></enumeration_b>
						<enumeration_c></enumeration_c>
						<chronology_i></chronology_i>
						<chronology_j></chronology_j> 
						<chronology_k></chronology_k>
						<public_note></public_note>
						<receiving_operator>Downiegl_API</receiving_operator>
						<internal_note_1></internal_note_1>
						<description></description>
						</item_data>
					</item>"""
				
				hold_stat = "<holding_id>{}</holding_id>".format(holding_id)
				barcode_stat = "<barcode>{}</barcode>".format(barcode)
				copy_stat = "<copy_id>{}</copy_id>".format(copy_id)
				# arrival_stat =  r'<arrival_date>{}</arrival_date>'.format(arrival_date)
				item_data = item_data.replace("<holding_id></holding_id>",hold_stat).replace("<barcode></barcode>", barcode_stat).replace("<copy_id></copy_id>", copy_stat)#.replace('<arrival_date></arrival_date>',arrival_stat)
				my_alma.create_item(mms_id, holding_id, item_data)
				if str(my_alma.status_code).startswith("2"):
					try:
						item_pid = re.findall(r"<pid>(.*?)</pid>",my_alma.xml_response_data)[0]
						sg.Print("Item created: ", item_pid)
						sg.Print("Item PID: ", item_pid)
						with open("item_report_"+arrival_date+".txt","a") as f:
							f.write(mms_id+"|"+holding_id+"|"+item_pid+"\n")
						return True
					except:
						return False
				else:
					print(my_alma.xml_response_data)
					try:
						error_message = re.findall(r"<errorMessage>(.*?)</errorMessage>", str(my_alma.xml_response_data))[0]
					except:
						error_message = "Failed: Something went wrong, but no error message is available from Alma API service"
					sg.Print(error_message)
					with open("item_report_"+arrival_date+".txt","a") as f:
						f.write(mms_id+"|"+holding_id+"| Failed \n")
					return False



def item_routine():
	"""
	This function manages the process of making items from a spreadsheet
	It manages GUI window closing, check if spreadsheet has necessary columns,
	taking holding id if it is in spreadsheet or set it None otherwise
	"""

	values = {}
	window = None

	while True:

		try:
			if window is None:
				values, window, event = my_gui(values)
			if event == 'Exit' or event == sg.WIN_CLOSED:
				window.close()
				sys.exit(0)
				break

			if event == "Run!":
				sg.Print("#"*50)
				spreadsheet_name = values['spreadsheet_name']
				sb = values['if_sb']
				my_alma_key = 'sb' if sb else 'prod'
				if sb:
					sg.Print("Making items in SANDBOX")
				else:
					sg.Print("Making items in PRODUCTION")
				if spreadsheet_name:
					if os.path.isfile(spreadsheet_name):
						wb = load_workbook(filename=spreadsheet_name, read_only=True)
						sheet = wb.active
						for row in sheet.iter_rows(min_row=2):
							mms_id = str(row[0].value).strip()
							if mms_id:
								barcode = str(row[1].value).strip()
								copy_id = str(row[2].value).strip()
								sg.Print(copy_id)
								location = str(row[3].value).strip()
								try:
									holding_id = str(row[4].value).strip()
								except:
									holding_id = None
								if mms_id and barcode:
									flag = make_item(mms_id, holding_id, location, copy_id, barcode, my_alma_key)
									print("flag", flag)

									if flag:
										sleep(5)

										sg.Print("Items created!")




								else:
									sg.PopupError("Spreadsheet does not have necessary columns.")
							event, values = window(timeout=0)
							if event == sg.WIN_CLOSED or event == 'Exit':
								break

					else:
						sg.PopupError("Selected file does not exist!")
				else:
					sg.PopupError("Please select a spreadsheet file!")

			window.close()
			window = None


		except Exception as e:
			
			print(e)

			if window:
				window.close()


			break


if __name__ == '__main__':

	item_routine()
